from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated, BasePermission
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Count, Avg, Max, Min
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.core.cache import cache
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import Processor, VideoCard, Motherboard, ComputerBuild, UserProfile
from .serializers import (
    ProcessorSerializer, VideoCardSerializer, MotherboardSerializer,
    ComputerBuildSerializer
)
import pyotp
import openpyxl
from openpyxl.styles import Font, Alignment

class OTPRequired(BasePermission):
    def has_permission(self, request, view):
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return True
        return cache.get(f'otp_good_{request.user.id}', False)

class ProcessorViewSet(viewsets.ModelViewSet):
    queryset = Processor.objects.all()
    serializer_class = ProcessorSerializer
    permission_classes = [IsAuthenticated, OTPRequired]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            owner = self.request.query_params.get('user_id')
            if owner:
                return Processor.objects.filter(user_id=owner)
            return Processor.objects.all()
        return Processor.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["GET"], url_path="stats")
    def stats(self, request):
        qs = self.get_queryset()
        stats = qs.aggregate(
            count=Count('id'),
            avg_cores=Avg('cores'),
            max_cores=Max('cores'),
            min_cores=Min('cores'),
            avg_frequency=Avg('frequency'),
            max_frequency=Max('frequency'),
            min_frequency=Min('frequency'),
        )
        return Response(stats)

    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request):
        qs = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Процессоры"
        
        headers = ['ID', 'Название', 'Ядра', 'Частота (ГГц)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, item in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.cores)
            ws.cell(row=row, column=4, value=float(item.frequency))
        
        for col in range(1, 5):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=processors.xlsx'
        wb.save(response)
        return response

class VideoCardViewSet(viewsets.ModelViewSet):
    queryset = VideoCard.objects.all()
    serializer_class = VideoCardSerializer
    permission_classes = [IsAuthenticated, OTPRequired]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            owner = self.request.query_params.get('user_id')
            if owner:
                return VideoCard.objects.filter(user_id=owner)
            return VideoCard.objects.all()
        return VideoCard.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["GET"], url_path="stats")
    def stats(self, request):
        qs = self.get_queryset()
        stats = qs.aggregate(
            count=Count('id'),
            avg_memory=Avg('memory'),
            max_memory=Max('memory'),
            min_memory=Min('memory'),
        )
        return Response(stats)

    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request):
        qs = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Видеокарты"
        
        headers = ['ID', 'Название', 'Память (ГБ)', 'Чипсет']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, item in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.memory)
            ws.cell(row=row, column=4, value=item.chipset)
        
        for col in range(1, 5):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=videocards.xlsx'
        wb.save(response)
        return response

class MotherboardViewSet(viewsets.ModelViewSet):
    queryset = Motherboard.objects.all()
    serializer_class = MotherboardSerializer
    permission_classes = [IsAuthenticated, OTPRequired]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            owner = self.request.query_params.get('user_id')
            if owner:
                return Motherboard.objects.filter(user_id=owner)
            return Motherboard.objects.all()
        return Motherboard.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["GET"], url_path="stats")
    def stats(self, request):
        qs = self.get_queryset()
        stats = qs.aggregate(
            count=Count('id'),
        )
        return Response(stats)

    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request):
        qs = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Материнские платы"
        
        headers = ['ID', 'Название', 'Сокет', 'Форм-фактор']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, item in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.socket)
            ws.cell(row=row, column=4, value=item.form_factor)
        
        for col in range(1, 5):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=motherboards.xlsx'
        wb.save(response)
        return response

class ComputerBuildViewSet(viewsets.ModelViewSet):
    queryset = ComputerBuild.objects.all()
    serializer_class = ComputerBuildSerializer
    permission_classes = [IsAuthenticated, OTPRequired]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            owner = self.request.query_params.get('user_id')
            if owner:
                return ComputerBuild.objects.filter(user_id=owner)
            return ComputerBuild.objects.all()
        return ComputerBuild.objects.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["GET"], url_path="stats")
    def stats(self, request):
        qs = self.get_queryset()
        stats = qs.aggregate(
            count=Count('id'),
            avg_price=Avg('price'),
            max_price=Max('price'),
            min_price=Min('price'),
        )
        return Response(stats)

    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request):
        qs = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сборки ПК"
        
        headers = ['ID', 'Название', 'Процессор', 'Видеокарта', 'Материнская плата', 'Цена']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, item in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.name)
            ws.cell(row=row, column=3, value=item.processor.name if item.processor else '')
            ws.cell(row=row, column=4, value=item.videocard.name if item.videocard else '')
            ws.cell(row=row, column=5, value=item.motherboard.name if item.motherboard else '')
            ws.cell(row=row, column=6, value=float(item.price))
        
        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=builds.xlsx'
        wb.save(response)
        return response

class UserViewSet(viewsets.GenericViewSet):
    queryset = User.objects.none()
    permission_classes = []

    @action(detail=False, url_path="info", methods=["GET"])
    def get_info(self, request):
        return Response({
            "username": request.user.username if request.user.is_authenticated else None,
            "is_authenticated": request.user.is_authenticated,
            "is_superuser": request.user.is_superuser if request.user.is_authenticated else False,
            "id": request.user.id
        })

    @action(detail=False, methods=["GET"], url_path="list_users")
    def list_users(self, request):
        if not request.user.is_superuser:
            return Response({"error": "Доступно только суперпользователям"}, status=403)
        users = User.objects.all().values('id', 'username')
        return Response(list(users))

    @method_decorator(csrf_exempt)
    @action(detail=False, url_path="login", methods=["POST"], permission_classes=[])
    def login_user(self, request):
        username = request.data.get("username")
        password = request.data.get("password")
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            otp_good = cache.get(f'otp_good_{user.id}', False)
            return Response({
                "success": True,
                "requires_otp": not otp_good,
                "is_authenticated": True
            })
        return Response({
            "success": False,
            "is_authenticated": False
        }, status=400)

    @method_decorator(csrf_exempt)
    @action(detail=False, url_path="logout", methods=["POST"])
    def logout_user(self, request):
        cache.delete(f'otp_good_{request.user.id}')
        logout(request)
        return Response({"success": True})

    @action(detail=False, url_path='otp-status', methods=['GET'])
    def get_otp_status(self, request, *args, **kwargs):
        otp_good = cache.get(f'otp_good_{request.user.id}', False)
        return Response({
            'otp_required': not otp_good
        })

    @action(detail=False, url_path='generate-otp', methods=['GET'])
    def generate_otp(self, request):
        if not request.user.is_authenticated:
            return Response({"error": "Требуется авторизация"}, status=401)
        
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        if not profile.otp_key:
            profile.otp_key = pyotp.random_base32()
            profile.save()
        
        return Response({
            "otp_key": profile.otp_key
        })

    @action(detail=False, url_path='verify-otp', methods=['POST'])
    def verify_otp(self, request):
        if not request.user.is_authenticated:
            return Response({"error": "Требуется авторизация"}, status=401)
        
        otp_code = request.data.get("otp_code")
        if not otp_code:
            return Response({"error": "OTP код обязателен"}, status=400)
        
        profile = get_object_or_404(UserProfile, user=request.user)
        if not profile.otp_key:
            return Response({"error": "OTP ключ не сгенерирован"}, status=400)
        
        totp = pyotp.TOTP(profile.otp_key)
        
        if totp.verify(otp_code):
            cache.set(f'otp_good_{request.user.id}', True, 60)
            return Response({"success": True, "message": "OTP подтверждён"})
        
        return Response({"success": False, "error": "Неверный OTP код"}, status=400)